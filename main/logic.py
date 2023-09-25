import openpyxl as xl
from openpyxl.styles import PatternFill, Alignment, Font
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from django.core.mail import send_mail
from email import encoders
from unhcr import settings
import time as tm
import threading
import smtplib

all_data = []
rsd_data = []
reg_data = []
residency_data = []
cbi_data = []
health_data = []
rst_data = []
protection_data = []


def email_sending(receiver, subject, filename):
    message = MIMEMultipart()
    message["to"] = receiver
    message["from"] = "Memezawy Inc."
    message["subject"] = subject

    attachment = open(filename, "rb")

    part = MIMEBase("application", 'vnd.ms-excel')
    part.set_payload(attachment.read())
    encoders.encode_base64(part)

    part.add_header('Content-Disposition', 'attachment', filename=filename)
    message.attach(part)

    with smtplib.SMTP(host="smtp.gmail.com", port=587) as smtp:
        smtp.ehlo()
        smtp.starttls()
        smtp.login("example@gmail.com", "password12")
        smtp.send_message(message)

    # resets data, so it does not duplicate in the next making


def data_reset():
    global all_data, rsd_data, rst_data, reg_data, cbi_data, health_data, protection_data, residency_data
    all_data = []
    rsd_data = []
    reg_data = []
    residency_data = []
    cbi_data = []
    health_data = []
    rst_data = []
    protection_data = []
    print('data has been reset')


subjects = ("RST", "RSD", "Health", "Protection", "CBI", "Residency", "Registration")
emails = ('RST@gmail.com', 'RSD@gmail.com',
          'Health@gmail.com', 'Protection@gmail.com', 'CBI@gmail.com',
          'Residency@gmail.com', 'Registration@gmail.com')


# Takes the data as a list of dictionaries and writes them in the worksheet
# Saves the data to their respective workbook to be sent later by email
def making_wb(wb_data, department):
    # creating the workbook
    wb = xl.Workbook()
    sheet = wb["Sheet"]

    # Styles the worksheet
    letters = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h']
    for letter in letters:
        sheet.column_dimensions[letter].width = 35

    for rows in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=8):
        for cell in rows:
            cell.fill = PatternFill(start_color='00FBFF', end_color='00FBFF', fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            fontStyle = Font(size="20", bold=True)
            cell.font = fontStyle

    # Adds the data types at the top
    data_types = ["name", "file_no", "phone_number", "nationality", "service", "service2", "message", "case_status"]
    style_data_types = ["Name", "File no.", "Phone no.", "Nationality", "Service 1", "Service 2", "Message"]
    data_col = 1
    for style_data_type in style_data_types:
        cell = sheet.cell(row=1, column=data_col)
        cell.value = style_data_type
        data_col += 1

    # assigns the data
    all_row = len(wb_data) + 1
    for col in range(1, 8):
        index = 0
        for row in range(2, all_row + 1):
            cell = sheet.cell(row=row, column=col)
            if data_types[col - 1] != 'case_status':
                cell.value = str(wb_data[index].get(data_types[col - 1]))
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                # sheet.row_dimensions[index].height = 45
                index += 1

    wb.save(f"{department}.xlsx")


receiver = ""
ind = 0


# this schedules the files to be sent at the specified time
def timer(datas, time):
    global ind
    for data in datas:

        inner_subject = subjects[ind]
        done = False

        while not done:

            dt = datetime.now()
            time_now = dt.strftime("%M")

            if time_now == time:
                making_wb(wb_data=data, department=inner_subject)
                print(inner_subject + " is subject made")

                filename = inner_subject + ".xlsx"

                email_sending(receiver=emails[ind], subject=f"{inner_subject} Department",
                              filename=filename)
                print(emails[ind])

                ind += 1
                done = True
    data_reset()


thread = threading.Thread()

i = True
# for reference
# subjects = ("RST", "RSD", "Health", "Protection", "CBI", "Residency", "Registration")

all_data = [rst_data, rsd_data, health_data, protection_data, cbi_data, residency_data, reg_data]


def receive_data(current_refugee, subject_received):
    global all_data, thread, i
    s = 0
    for subject in subjects:
        if subject_received == subjects[s]:
            all_data[s].append(current_refugee)
        s += 1
    thread = threading.Thread(target=timer, args=(all_data, "01"))
    if i:
        thread.start()
        i = False
